import xlrd
import openpyxl 
from ..models import Categoria, Produto


def import_xlsx(filename):
    workbook = openpyxl.load_workbook(filename) 
    sheet = workbook.active 
    categorias = []
    for row_idx in range(2, sheet.max_row + 1): 
        categoria_cell = sheet.cell(row=row_idx, column=7) 
        categoria = categoria_cell.value
        if categoria: 
            categorias.append(categoria)

    categorias_unicas = [Categoria(categoria=categoria)
                        for categoria in set(categorias) if categoria]

    Categoria.objects.all().delete()
    Categoria.objects.bulk_create(categorias_unicas)

    aux = []
    for row_idx in range(2, sheet.max_row + 1):
        produto = sheet.cell(row=row_idx, column=1).value
        ncm = sheet.cell(row=row_idx, column=2).value
        try:
            ncm = int(ncm) 
        except (ValueError, TypeError):
            ncm = None 
        _importado = sheet.cell(row=row_idx, column=3).value 
        importado = True if str(_importado).lower() == 'true' else False 
        preco = sheet.cell(row=row_idx, column=4).value 
        estoque = sheet.cell(row=row_idx, column=5).value 
        estoque_minimo = sheet.cell(row=row_idx, column=6).value 

        _categoria_nome = sheet.cell(row=row_idx, column=7).value 
        categoria = Categoria.objects.filter(categoria=_categoria_nome).first()

        produto_data = dict(
            produto=produto,
            ncm=ncm,
            importado=importado,
            preco=preco,
            estoque=estoque,
            estoque_minimo=estoque_minimo,
        )

        if categoria:
            obj = Produto(categoria=categoria, **produto_data)
        else:
            obj = Produto(**produto_data)

        aux.append(obj)

    Produto.objects.all().delete()
    Produto.objects.bulk_create(aux)